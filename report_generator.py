"""
第八步：报告生成模块
一键生成Excel/PDF分析报告
"""

import pandas as pd
import sqlite3
import os
from datetime import datetime
import json

# 尝试导入Excel相关库
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl 未安装，Excel报告功能不可用")

db_path = os.path.join(os.path.dirname(__file__), 'stock_data.db')
REPORT_DIR = 'reports'

if not os.path.exists(REPORT_DIR):
    os.makedirs(REPORT_DIR)


def generate_excel_report(symbol, name, output_file=None):
    """
    生成单只股票的Excel分析报告
    
    参数:
        symbol: 股票代码
        name: 股票名称
        output_file: 输出文件名（不指定则自动生成）
    """
    if not EXCEL_AVAILABLE:
        print("❌ Excel功能不可用，请安装: pip install openpyxl")
        return False
    
    try:
        conn = sqlite3.connect(db_path)
        
        # 获取K线数据
        table_name = f'kline_stock_{symbol}'
        try:
            df = pd.read_sql(f"SELECT * FROM {table_name} ORDER BY trade_date DESC LIMIT 60", conn)
        except:
            df = pd.DataFrame()
        
        # 获取实时数据
        try:
            realtime = pd.read_sql(f"SELECT * FROM stock_realtime WHERE 代码='{symbol}'", conn)
        except:
            realtime = pd.DataFrame()
        
        conn.close()
        
        if df.empty:
            print(f"❌ {name} 没有K线数据")
            return False
        
        # 生成输出文件名
        if output_file is None:
            today = datetime.now().strftime('%Y%m%d')
            output_file = os.path.join(REPORT_DIR, f'{name}_{symbol}_{today}.xlsx')
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        
        # 1. 概述页
        ws_summary = wb.active
        ws_summary.title = '概述'
        
        # 标题
        ws_summary['A1'] = f'{name} ({symbol}) 分析报告'
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary.merge_cells('A1:F1')
        
        # 基本信息
        ws_summary['A3'] = '报告时间:'
        ws_summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if not realtime.empty:
            row = realtime.iloc[0]
            ws_summary['A4'] = '最新价格:'
            ws_summary['B4'] = row.get('最新价', 'N/A')
            ws_summary['A5'] = '涨跌幅:'
            ws_summary['B5'] = f"{row.get('涨跌幅', 0):.2f}%"
            ws_summary['A6'] = '成交量:'
            ws_summary['B6'] = row.get('成交量', 'N/A')
            ws_summary['A7'] = '成交额:'
            ws_summary['B7'] = row.get('成交额', 'N/A')
        
        # 2. 行情数据页
        ws_data = wb.create_sheet('行情数据')
        headers = ['日期', '开盘', '最高', '最低', '收盘', '成交量']
        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # 写入数据
        for row_idx, (_, row) in enumerate(df.head(30).iterrows(), 2):
            ws_data.cell(row_idx, 1, row.get('trade_date', ''))
            ws_data.cell(row_idx, 2, row.get('open', 0))
            ws_data.cell(row_idx, 3, row.get('high', 0))
            ws_data.cell(row_idx, 4, row.get('low', 0))
            ws_data.cell(row_idx, 5, row.get('close', 0))
            ws_data.cell(row_idx, 6, row.get('volume', 0))
        
        # 设置列宽
        for col in range(1, 7):
            ws_data.column_dimensions[chr(64 + col)].width = 12
        
        # 3. 统计信息页
        ws_stats = wb.create_sheet('统计信息')
        ws_stats['A1'] = '统计指标'
        ws_stats['A1'].font = Font(bold=True)
        ws_stats['B1'] = '数值'
        ws_stats['B1'].font = Font(bold=True)
        
        stats = [
            ('数据总数', len(df)),
            ('最高价', df['high'].max() if not df.empty else 0),
            ('最低价', df['low'].min() if not df.empty else 0),
            ('平均价', df['close'].mean() if not df.empty else 0),
            ('区间涨跌幅', f"{((df.iloc[-1]['close'] - df.iloc[0]['close']) / df.iloc[0]['close'] * 100):.2f}%")
        ]
        
        for idx, (label, value) in enumerate(stats, 2):
            ws_stats.cell(idx, 1, label)
            ws_stats.cell(idx, 2, value)
        
        # 保存文件
        wb.save(output_file)
        print(f"✅ 报告已生成: {output_file}")
        return output_file
        
    except Exception as e:
        print(f"❌ 生成报告失败: {e}")
        return False


def generate_portfolio_report(stock_list, output_file=None):
    """
    生成组合报告（多只股票对比）
    
    参数:
        stock_list: [(code, name), ...] 股票列表
        output_file: 输出文件名
    """
    if not EXCEL_AVAILABLE:
        print("❌ Excel功能不可用，请安装: pip install openpyxl")
        return False
    
    try:
        conn = sqlite3.connect(db_path)
        
        # 获取实时数据
        codes = [s[0] for s in stock_list]
        placeholders = ','.join(['?' for _ in codes])
        
        try:
            df = pd.read_sql(
                f"SELECT * FROM stock_realtime WHERE 代码 IN ({placeholders})",
                conn,
                params=codes
            )
        except:
            df = pd.DataFrame()
        
        conn.close()
        
        if df.empty:
            print("❌ 没有找到这些股票的数据")
            return False
        
        # 生成输出文件名
        if output_file is None:
            today = datetime.now().strftime('%Y%m%d')
            output_file = os.path.join(REPORT_DIR, f'组合报告_{today}.xlsx')
        
        # 创建Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '组合概览'
        
        # 标题
        ws['A1'] = f'组合报告 ({datetime.now().strftime("%Y-%m-%d %H:%M")})'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:H1')
        
        # 表头
        headers = ['代码', '名称', '最新价', '涨跌幅', '成交量', '成交额', '最高', '最低']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # 写入数据
        for row_idx, (_, row) in enumerate(df.iterrows(), 3):
            ws.cell(row_idx, 1, row.get('代码', ''))
            ws.cell(row_idx, 2, row.get('名称', ''))
            ws.cell(row_idx, 3, row.get('最新价', 0))
            ws.cell(row_idx, 4, row.get('涨跌幅', 0))
            ws.cell(row_idx, 5, row.get('成交量', 0))
            ws.cell(row_idx, 6, row.get('成交额', 0))
            ws.cell(row_idx, 7, row.get('最高', 0))
            ws.cell(row_idx, 8, row.get('最低', 0))
        
        # 设置列宽
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 12
        
        wb.save(output_file)
        print(f"✅ 组合报告已生成: {output_file}")
        return output_file
        
    except Exception as e:
        print(f"❌ 生成组合报告失败: {e}")
        return False


# 测试函数
if __name__ == "__main__":
    print("=" * 50)
    print("📄 测试报告生成模块")
    print("=" * 50)
    
    # 生成单只股票报告
    generate_excel_report('000001', '平安银行')
    
    # 生成组合报告
    generate_portfolio_report([('000001', '平安银行'), ('600519', '贵州茅台')])
    
    print("\n✅ 报告生成模块测试完成！")